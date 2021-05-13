import logging
import argparse
import lxml.etree as etree
import sys
from openpyxl import load_workbook
from openpyxl import utils

def workbook_to_tree(workbook, mode):
    """Takes an openpyxl workbook object and transforms it to an XML.
       It can be specified how the XML tags should be created:
       default: tagnames are workbook/sheet/row/column all tags have an index attribute.
       short_tags: tagnames are w/s/r/c index attribute is i (to reduce size in big files)
       tags_from_file: column tags will be take from the first row (in case they are XML tag
                       compatible strings. Otherwise default tag names will be used.)
    """
    if mode == 'short_tags':
        workbook_tag_name = 'w'
        sheet_tag_name = 's'
        row_tag_name = 'r'
        column_tag_name = 'c'
        index_tag_name = 'i'
    else:
        workbook_tag_name = 'workbook'
        sheet_tag_name = 'sheet'
        row_tag_name = 'row'
        column_tag_name = 'column'
        index_tag_name = 'index'

    root = etree.Element(workbook_tag_name)
    for i, sheetname in enumerate(workbook.sheetnames):

        sheet = workbook[sheetname]
        if mode == 'tags_from_file':
            if valid_XML_tag(sheetname):
                sheet_tag_name = sheetname
            else: sheet_tag_name = 'sheet'
            first_row = sheet[1]
        sheet_tag = etree.SubElement(root, sheet_tag_name)
        sheet_tag.set(index_tag_name, str(i + 1))

        for j, row in enumerate(sheet.iter_rows(values_only=True)):
            if mode == 'tags_from_file' and j == 0:
                pass
            else:
                row_tag = etree.SubElement(sheet_tag, row_tag_name)
                row_tag.set(index_tag_name, str(j + 1))
                for k, column in enumerate(row):
                    if mode == 'tags_from_file' and valid_XML_tag(first_row[k].value):
                        column_tag_name = first_row[k].value
                    else: column_tag_name = 'column'
                    column_tag = etree.SubElement(row_tag, column_tag_name)
                    column_tag.text = str(column) if column != None else ""
                    column_tag.set(index_tag_name, str(k + 1))

    tree = etree.ElementTree(root)
    return tree

def valid_XML_tag(str):
    """Checks if a string can be used as a XML tag."""
    not_allowed_in_xml_tags = '!"\'#$%&()*+,/;<=>?@[\]^`{|}~ '
    return True not in [c in str for c in not_allowed_in_xml_tags] and \
                str[0] not in ("-", ".", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0")

def main():
    """Used when executed from command line"""

    h_verbose = 'Increase output verbosity'
    h_input   = 'Path and name to the input file that you want to transform.'
    h_output  = 'Path and name to the output xml file that you want to create or \
                "-" for stdout.'
    h_pretty  = 'Pretty print output.'
    h_mode    = 'Which xml tags should be used in the output xml?'

    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "--verbose", help=h_verbose, action="store_true")
    parser.add_argument("-i", "--input", help=h_input, nargs=1, required=True)
    parser.add_argument("-o", "--output", help=h_output, default='-', nargs=1)
    parser.add_argument("-fo", "--pretty", help=h_pretty, action="store_true")
    parser.add_argument("-m", "--mode", help=h_mode, \
                        choices=['default', 'short_tags', 'tags_from_file'], \
                        default='default')

    args = parser.parse_args()
    if args.verbose:
        loglevel = 'INFO'
    else:
        loglevel = 'WARNING'

    if args.input:
        input_xlsx_name = args.input[0]
    if args.output:
        output_file_name = args.output[0]

    pretty_print = True if args.pretty else False

    mode = args.mode

    numeric_level = getattr(logging, loglevel.upper())
    logging.basicConfig(level=numeric_level)

    try:
        workbook = load_workbook(filename=input_xlsx_name)

    except FileNotFoundError as error:
        print('Error: {}'.format(error))
        sys.exit(3)
    except utils.exceptions.InvalidFileException as error:
        print('Error: {}'.format(error))
        sys.exit(2)
    except:
        print('Error {}'.format(sys.exc_info()))
        sys.exit(1)

    tree = workbook_to_tree(workbook, mode)

    if not args.output or output_file_name == '-' :
        sys.stdout.write(etree.tostring(tree, pretty_print=pretty_print, encoding='unicode'))
    elif len(output_file_name) > 0:
        try:
            tree.write(output_file_name, pretty_print=pretty_print, encoding='UTF-8')
        except PermissionError as error:
            print('Failed to write to file: {}'.format(error))
            quit()

if __name__ == "__main__":
    main()
